from __future__ import annotations

import re
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from tools.document_reader.models import Document, ReaderLimits

NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
}


def parse_xlsx(path: Path, limits: ReaderLimits) -> Document:
    try:
        import openpyxl  # type: ignore
    except ImportError:
        return _parse_xlsx_zip(path, limits)

    warnings: list[str] = []
    try:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        return Document(path.name, "xlsx", warnings=[f"Could not parse workbook: {exc}"])

    sheets = []
    for sheet_index, ws in enumerate(workbook.worksheets):
        if sheet_index >= limits.max_sheets:
            warnings.append(f"Sheets truncated to {limits.max_sheets}.")
            break
        max_row = ws.max_row or 0
        max_column = ws.max_column or 0
        rows = []
        formulas = []
        row_limit = min(max_row, limits.max_rows)
        for row in ws.iter_rows(min_row=1, max_row=row_limit, max_col=max_column):
            values = []
            for cell in row:
                value = cell.value
                values.append(value)
                if isinstance(value, str) and value.startswith("="):
                    formulas.append({"cell": cell.coordinate, "formula": value})
            rows.append(values)
        if max_row > limits.max_rows:
            warnings.append(f"Sheet {ws.title} rows truncated to {limits.max_rows} from {max_row}.")
        headers = rows[0] if rows else []
        sheets.append(
            {
                "name": ws.title,
                "state": ws.sheet_state,
                "used_range": ws.calculate_dimension(),
                "headers": headers,
                "rows": rows[1:],
                "formulas": formulas,
                "merged_cells": [str(rng) for rng in ws.merged_cells.ranges],
                "tables": [
                    {"name": name, "range": table.ref}
                    for name, table in getattr(ws, "tables", {}).items()
                ],
            }
        )
    named_ranges = []
    try:
        for defined_name in workbook.defined_names.values():
            named_ranges.append({"name": defined_name.name, "text": defined_name.attr_text})
    except Exception:
        warnings.append("Could not extract named ranges.")
    if not sheets:
        warnings.append("Workbook has no visible worksheet content.")
    return Document(
        file_name=path.name,
        file_type="xlsx",
        metadata={"workbook": path.name, "sheet_names": workbook.sheetnames, "named_ranges": named_ranges},
        sheets=sheets,
        warnings=warnings,
    )


def _parse_xlsx_zip(path: Path, limits: ReaderLimits) -> Document:
    warnings = ["openpyxl is not installed; using limited XLSX XML parser."]
    try:
        with zipfile.ZipFile(path) as archive:
            shared_strings = _read_shared_strings(archive)
            rels = _read_workbook_rels(archive)
            workbook_xml = ET.fromstring(archive.read("xl/workbook.xml"))
            sheet_defs = workbook_xml.findall(".//main:sheets/main:sheet", NS)
            sheets = []
            for index, sheet in enumerate(sheet_defs):
                if index >= limits.max_sheets:
                    warnings.append(f"Sheets truncated to {limits.max_sheets}.")
                    break
                name = sheet.attrib.get("name", f"Sheet{index + 1}")
                state = sheet.attrib.get("state", "visible")
                rel_id = sheet.attrib.get(f"{{{NS['rel']}}}id")
                target = rels.get(rel_id or "", f"worksheets/sheet{index + 1}.xml")
                worksheet_path = "xl/" + target.lstrip("/")
                if worksheet_path not in archive.namelist():
                    warnings.append(f"Could not find worksheet XML for {name}.")
                    continue
                sheets.append(_parse_sheet_xml(archive, worksheet_path, name, state, shared_strings, limits, warnings))
            return Document(
                file_name=path.name,
                file_type="xlsx",
                metadata={"workbook": path.name, "sheet_names": [s.attrib.get("name") for s in sheet_defs]},
                sheets=sheets,
                warnings=warnings,
            )
    except zipfile.BadZipFile:
        return Document(path.name, "xlsx", warnings=["Could not parse workbook: invalid or corrupt XLSX archive."])
    except Exception as exc:
        return Document(path.name, "xlsx", warnings=[f"Could not parse workbook: {exc}"])


def _read_shared_strings(archive: zipfile.ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root.findall(".//main:si", NS):
        strings.append("".join(text.text or "" for text in item.findall(".//main:t", NS)))
    return strings


def _read_workbook_rels(archive: zipfile.ZipFile) -> dict[str, str]:
    if "xl/_rels/workbook.xml.rels" not in archive.namelist():
        return {}
    root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rels = {}
    for rel in root:
        rels[rel.attrib.get("Id", "")] = rel.attrib.get("Target", "")
    return rels


def _parse_sheet_xml(
    archive: zipfile.ZipFile,
    worksheet_path: str,
    name: str,
    state: str,
    shared_strings: list[str],
    limits: ReaderLimits,
    warnings: list[str],
) -> dict:
    root = ET.fromstring(archive.read(worksheet_path))
    dimension = root.find("main:dimension", NS)
    used_range = dimension.attrib.get("ref", "") if dimension is not None else ""
    rows = []
    formulas = []
    for row_index, row in enumerate(root.findall(".//main:sheetData/main:row", NS)):
        if row_index >= limits.max_rows:
            warnings.append(f"Sheet {name} rows truncated to {limits.max_rows}.")
            break
        values = []
        for cell in row.findall("main:c", NS):
            value, formula = _cell_value(cell, shared_strings)
            values.append(value)
            if formula is not None:
                formulas.append({"cell": cell.attrib.get("r"), "formula": formula})
        rows.append(values)
    merged_cells = [
        item.attrib.get("ref")
        for item in root.findall(".//main:mergeCells/main:mergeCell", NS)
        if item.attrib.get("ref")
    ]
    headers = rows[0] if rows else []
    return {
        "name": name,
        "state": state,
        "used_range": used_range,
        "headers": headers,
        "rows": rows[1:],
        "formulas": formulas,
        "merged_cells": merged_cells,
        "tables": [],
    }


def _cell_value(cell: ET.Element, shared_strings: list[str]) -> tuple[str | float | int | None, str | None]:
    formula_element = cell.find("main:f", NS)
    formula = f"={formula_element.text}" if formula_element is not None and formula_element.text else None
    value_element = cell.find("main:v", NS)
    value = value_element.text if value_element is not None else None
    if value is not None and cell.attrib.get("t") == "s":
        try:
            value = shared_strings[int(value)]
        except (ValueError, IndexError):
            pass
    if isinstance(value, str) and re.fullmatch(r"-?\d+", value):
        return int(value), formula
    if isinstance(value, str) and re.fullmatch(r"-?\d+\.\d+", value):
        return float(value), formula
    return value, formula

